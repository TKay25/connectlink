import pandas as pd
import numpy as np
from tkinter import Tk, filedialog
from fuzzywuzzy import fuzz, process
import openpyxl
from openpyxl.styles import PatternFill
import os

def select_file(file_description):
    """Open file dialog to select an Excel file"""
    Tk().withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(
        title=f"Select {file_description} Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    return file_path

def highlight_similar_customers(loans_df, loc_df, loc_name, threshold=80):
    """
    Highlight rows in loans_df where customer name is similar to any name in loc_df
    threshold: minimum similarity score (0-100)
    """
    # Convert to string and handle NaN values
    loans_names = loans_df['CUSTOMER_NAME1'].astype(str).fillna('')
    loc_names = loc_df['Name of Client'].astype(str).fillna('').tolist()
    
    # Create a boolean mask for matches
    matches = []
    similarity_scores = []
    
    for loan_name in loans_names:
        if loan_name and loan_name != 'nan':
            # Find best match from LOC names
            best_match, score = process.extractOne(loan_name, loc_names, scorer=fuzz.token_sort_ratio)
            if score >= threshold:
                matches.append(True)
                similarity_scores.append(score)
            else:
                matches.append(False)
                similarity_scores.append(0)
        else:
            matches.append(False)
            similarity_scores.append(0)
    
    # Add similarity score column for reference
    loans_df[f'Similarity_Score_{loc_name}'] = similarity_scores
    
    return matches

def apply_highlighting_to_excel(file_path, loans_df, highlight_colors):
    """
    Apply highlighting to the original Excel file
    """
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Define fill colors
    fills = {
        'LOC1': PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid"),  # Light orange
        'LOC2': PatternFill(start_color="B4FFE5", end_color="B4FFE5", fill_type="solid")   # Light mint
    }
    
    # Find the column index for CUSTOMER_NAME1 (assuming it's in row 1)
    header_row = 1
    customer_col_idx = None
    
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value == 'CUSTOMER_NAME1':
            customer_col_idx = col_idx
            break
    
    if customer_col_idx is None:
        print("Warning: 'CUSTOMER_NAME1' column not found in the Excel file")
        return
    
    # Apply highlighting based on matches
    for row_idx, (loc1_match, loc2_match) in enumerate(zip(highlight_colors['LOC1'], highlight_colors['LOC2']), start=2):
        if loc1_match:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fills['LOC1']
        elif loc2_match:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fills['LOC2']
    
    # Save the highlighted file with a new name
    base_name = os.path.splitext(file_path)[0]
    new_file_path = f"{base_name}_highlighted.xlsx"
    wb.save(new_file_path)
    print(f"Highlighted file saved as: {new_file_path}")
    
    return new_file_path

def main():
    print("Please select the 3 Excel files when prompted...")
    
    # Select the three files
    loans_file = select_file("LOANS")
    if not loans_file:
        print("No file selected for LOANS. Exiting.")
        return
    
    loc1_file = select_file("LOC1")
    if not loc1_file:
        print("No file selected for LOC1. Exiting.")
        return
    
    loc2_file = select_file("LOC2")
    if not loc2_file:
        print("No file selected for LOC2. Exiting.")
        return
    
    print("\nLoading files...")
    
    # Load the data
    try:
        loans_df = pd.read_excel(loans_file)
        loc1_df = pd.read_excel(loc1_file)
        loc2_df = pd.read_excel(loc2_file)
        
        print(f"Loans file loaded: {len(loans_df)} rows")
        print(f"LOC1 file loaded: {len(loc1_df)} rows")
        print(f"LOC2 file loaded: {len(loc2_df)} rows")
        
        # Check if required columns exist
        if 'CUSTOMER_NAME1' not in loans_df.columns:
            print("Error: 'CUSTOMER_NAME1' column not found in Loans file!")
            print(f"Available columns: {loans_df.columns.tolist()}")
            return
        
        if 'Name of Client' not in loc1_df.columns:
            print("Error: 'Name of Client' column not found in LOC1 file!")
            print(f"Available columns: {loc1_df.columns.tolist()}")
            return
        
        if 'Name of Client' not in loc2_df.columns:
            print("Error: 'Name of Client' column not found in LOC2 file!")
            print(f"Available columns: {loc2_df.columns.tolist()}")
            return
        
    except Exception as e:
        print(f"Error loading files: {e}")
        return
    
    # Set similarity threshold (adjust as needed)
    SIMILARITY_THRESHOLD = 80  # 70% similarity threshold
    
    print(f"\nMatching customers with {SIMILARITY_THRESHOLD}% similarity threshold...")
    
    # Find similar customers
    loc1_matches = highlight_similar_customers(loans_df, loc1_df, 'LOC1', SIMILARITY_THRESHOLD)
    loc2_matches = highlight_similar_customers(loans_df, loc2_df, 'LOC2', SIMILARITY_THRESHOLD)
    
    # Create a summary DataFrame
    loans_df['Match_with_LOC1'] = loc1_matches
    loans_df['Match_with_LOC2'] = loc2_matches
    
    # Count matches
    loc1_match_count = sum(loc1_matches)
    loc2_match_count = sum(loc2_matches)
    both_match_count = sum([m1 and m2 for m1, m2 in zip(loc1_matches, loc2_matches)])
    
    print(f"\nResults:")
    print(f"- Matches with LOC1: {loc1_match_count} rows")
    print(f"- Matches with LOC2: {loc2_match_count} rows")
    print(f"- Matches with both: {both_match_count} rows")
    
    # Show sample of matches
    matched_loans = loans_df[loans_df['Match_with_LOC1'] | loans_df['Match_with_LOC2']]
    if len(matched_loans) > 0:
        print(f"\nSample of matched customers:")
        print(matched_loans[['CUSTOMER_NAME1', 'Similarity_Score_LOC1', 'Similarity_Score_LOC2']].head(10))
    
    # Save the results to a new Excel file with summary
    output_file = os.path.join(os.path.dirname(loans_file), "loans_with_matches.xlsx")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write the detailed data
        loans_df.to_excel(writer, sheet_name='Loans with Matches', index=False)
        
        # Write summary
        summary_data = {
            'Metric': ['Total Loans', 'Matches with LOC1', 'Matches with LOC2', 'Matches with Both'],
            'Count': [len(loans_df), loc1_match_count, loc2_match_count, both_match_count],
            'Percentage': [
                '100%',
                f"{(loc1_match_count/len(loans_df)*100):.1f}%",
                f"{(loc2_match_count/len(loans_df)*100):.1f}%",
                f"{(both_match_count/len(loans_df)*100):.1f}%"
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    print(f"\nDetailed results saved to: {output_file}")
    
    # Apply highlighting to the original Loans file
    print("\nCreating highlighted Excel file...")
    highlight_colors = {
        'LOC1': loc1_matches,
        'LOC2': loc2_matches
    }
    highlighted_file = apply_highlighting_to_excel(loans_file, loans_df, highlight_colors)
    
    print("\nProcess completed successfully!")
    print(f"1. Detailed results with similarity scores: {output_file}")
    if highlighted_file:
        print(f"2. Highlighted original file: {highlighted_file}")
    
    # Display the first few matched records
    print("\nFirst 10 matched records (showing similarity scores):")
    matched_display = matched_loans[['CUSTOMER_NAME1', 'Similarity_Score_LOC1', 'Similarity_Score_LOC2']].head(10)
    print(matched_display.to_string(index=False))

if __name__ == "__main__":
    # Install required packages if not already installed
    required_packages = ['pandas', 'openpyxl', 'fuzzywuzzy', 'python-Levenshtein', 'tkinter']
    
    try:
        main()
    except ImportError as e:
        print(f"Missing required package: {e}")
        print("\nPlease install required packages using:")
        print("pip install pandas openpyxl fuzzywuzzy python-Levenshtein")
        print("\nNote: tkinter usually comes with Python by default")